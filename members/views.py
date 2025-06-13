from django.shortcuts import render
from django.views import View
from django.core.paginator import Paginator
from django.http import HttpResponse
from openpyxl.workbook import Workbook

from .admin_login_view import AdminRequiredMixin
from .forms import AuthorForm, BookForm, BorrowRecordForm
from .models import Author, Book, BorrowRecord


class AuthorViews(AdminRequiredMixin,View):
    def get(self, request):
        form = AuthorForm()
        return render(request, 'Author.html', {'form': form})

    def post(self, request):
        form = AuthorForm(request.POST)
        if form.is_valid():
            form.save()
            return render(request, 'Author.html', {'form': AuthorForm(), 'success': 'Author added'})
        return render(request, 'Author.html', {'form': form})


class BookViews(AdminRequiredMixin,View):
    def get(self, request):
        form = BookForm()
        return render(request, 'Book.html', {'form': form})

    def post(self, request):
        form = BookForm(request.POST)
        if form.is_valid():
            form.save()
            return render(request, 'Book.html', {'form': BookForm(), 'success': 'Book added'})
        return render(request, 'Book.html', {'form': form})


class BorrowRecordViews(AdminRequiredMixin,View):
    def get(self, request):
        form = BorrowRecordForm()
        return render(request, 'BorrowRecord.html', {'form': form})

    def post(self, request):
        form = BorrowRecordForm(request.POST)
        if form.is_valid():
            form.save()
            return render(request, 'BorrowRecord.html', {'form': BorrowRecordForm(), 'success': 'Borrow record added'})
        return render(request, 'BorrowRecord.html', {'form': form})



class AuthorListView(View):
    def get(self, request):
        authors = Author.objects.all()
        paginator = Paginator(authors, 5) 
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        return render(request, 'AuthorList.html', {'page_obj': page_obj})


class BookListView(View):
    def get(self, request):
        books = Book.objects.all()
        paginator = Paginator(books, 5)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        return render(request, 'BookList.html', {'page_obj': page_obj})


class BorrowRecordListView(View):
    def get(self, request):
        records = BorrowRecord.objects.all()
        paginator = Paginator(records, 5)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        return render(request, 'BorrowRecordList.html', {'page_obj': page_obj})



class ExportExcelView(AdminRequiredMixin,View):
    def get(self, request):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Authors"
        ws1.append(["Name", "Email", "Bio"])
        for author in Author.objects.all():
            ws1.append([author.name, author.email, author.bio])

        
        ws2 = wb.create_sheet(title="Books")
        ws2.append(["Title", "Genre", "Published Date", "Author"])
        for book in Book.objects.all():
            ws2.append([book.title, book.genre, str(book.published_date), book.author.name])
        ws3 = wb.create_sheet(title="BorrowRecords")
        ws3.append(["User Name", "Book", "Borrow Date", "Return Date"])
        for record in BorrowRecord.objects.all():
            ws3.append([record.user_name, record.book.title, str(record.borrow_date), str(record.return_date)])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=library_data.xlsx'
        wb.save(response)
        return response
